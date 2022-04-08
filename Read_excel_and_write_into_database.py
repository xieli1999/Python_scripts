#pip3 install xlrd



#%pip install openpyxl
#%pip install mysql
#%pip install sqlalchemy
#pip3 install PyMySql
#pip3 install cloud-sql-python-connector[pymysql]

import openpyxl
import pathlib
import os
import pymysql
from google.cloud.sql.connector import connector
import sqlalchemy
import datetime
import time
import threading
import math

from datetime import datetime
from datetime import timedelta
from datetime import date
from re import M, S

#the following modules are my own modules
import Excel_Functions
import Class_Definition
import MySQL_Functions

from Class_Definition import Asset, Spouse, Beneficiary, Employment, Liability, Person, ResponseDTO, Advisor, Trustee, Address, ID, Liability, KYC, Income
from Excel_Functions import CN_Prov_of_birth, Validate_SIN, Date_of_Birth
from MySQL_Functions import AIFinanceDB

def main_application(entries):
    global PersonID_index
    db=AIFinanceDB()
    
    for entry in entries:
        path=my_excels_dir.joinpath(entry)
        workbook=openpyxl.load_workbook(path, data_only = True)
        sheet=workbook['Account Info']

        #D_of_application = sheet['F3'].value
        #M_of_application = sheet['D3'].value
        #Y_of_application = sheet['B3'].value
        #Application_date = Date_of_Birth(Y_of_application,M_of_application,D_of_application)    
        Application_date = date.today() 

        sheet=workbook['Personal Info']
        Primary_verification_date = sheet['J18'].value
        Primary_verification_Notes = sheet['H19'].value

        
        Primary_applicant = Person()
        Primary_applicant.PersonID = create_personID()
        Primary_applicant.First_Name = sheet['E7'].value
        Primary_applicant.Last_Name=sheet['B7'].value
        
        if Primary_applicant.First_Name != None and Primary_applicant.Last_Name != None:
            Primary_applicant.English_Name=sheet['J7'].value
            Primary_applicant.Gender=sheet['B8'].value
            
            D_of_Birth=sheet['H8'].value
            M_of_Birth=sheet['F8'].value
            Y_of_Birth=sheet['D8'].value
            if Y_of_Birth != None and M_of_Birth!= None and D_of_Birth!= None: # Do nothing if there is no Date of Birth
                Primary_applicant.Date_of_Birth=Date_of_Birth(Y_of_Birth,M_of_Birth,D_of_Birth)
            
                place_of_birth = sheet['E10'].value
                if place_of_birth is None:
                    Primary_applicant.Country_of_Birth = ""
                    Primary_applicant.Province_of_Birth = ""
                else:
                    Primary_applicant.Country_of_Birth=CN_Prov_of_birth(place_of_birth)[0]
                    Primary_applicant.Province_of_Birth=CN_Prov_of_birth(place_of_birth)[1]

                Primary_applicant.Citizenship=sheet['G9'].value
                Primary_applicant.Tax_Status=sheet['K9'].value
                
                if sheet['I10'].value == "Birth":
                    Primary_applicant.Live_in_Canada_Since=Primary_applicant.Date_of_Birth
                elif sheet['I10'].value == "Date (mm/yyyy)" :
                    Primary_applicant.Live_in_Canada_Since=sheet['L10'].value
                else: Primary_applicant.Live_in_Canada_Since = None
                
                Primary_applicant.Marital_Status=sheet['D9'].value
                Primary_applicant.Cellphone=sheet['F16'].value
                Primary_applicant.Homephone=sheet['H16'].value
                Primary_applicant.Workphone=sheet['K16'].value

                SINNo = list(str(sheet['K8'].value))
                if len(SINNo) == 9:
                    if Validate_SIN(SINNo):
                        Primary_applicant.SIN = str(sheet['K8'].value)
                    else:
                        print(Primary_applicant.Full_Name + ' ' + str(sheet['K8'].value) + ': The SIN is NOT valid')
                        Primary_applicant.SIN = str(sheet['K8'].value)    ##for test purpose only  ''  
                else:
                    Primary_applicant.SIN = None
                
                Primary_applicant.Email=sheet['A16'].value
                Primary_applicant.Postcode=sheet['I12'].value
                Primary_applicant.PersonType="" 

                if sheet['E20'].value == "Yes":
                    Primary_applicant.Bankruptcy = 1
                    Primary_applicant.Discharge_Date = sheet['J20'].value
                else:
                    Primary_applicant.Bankruptcy = 0
                    Primary_applicant.Discharge_Date = None
                
                Primary_applicant.AdvisorPID = sheet['B3'].value
                if Primary_applicant.AdvisorPID == None or Primary_applicant.AdvisorPID == 'None':
                    Primary_applicant.AdvisorPID = 'Virtual Advisor' 
                Primary_applicant.Start_date = None 
                Primary_applicant.End_date = None
                Primary_applicant.Current_Flag = 1
                Primary_applicant.Verify_date = Primary_verification_date
                Primary_applicant.Notes = Primary_verification_Notes
                    
                result=db.SaveApplicant(Primary_applicant)
                
                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                
                # The end of Primary applicant

                # Save current address

                if sheet['B12'].value != None:
                    Primary_address = Address()

                    Primary_address.PersonID = Primary_applicant.PersonID
                    Primary_address.Apt_No = str(sheet['E12'].value)
                    Primary_address.Street_No = str(sheet['A12'].value)
                    Primary_address.Street_Name = sheet['B12'].value
                    Primary_address.City = sheet['F12'].value
                    Primary_address.Province = sheet['H12'].value
                    Primary_address.Country = "Canada"
                    Primary_address.Postcode = sheet['I12'].value
                    Primary_address.Living_Status = sheet['B9'].value
                    
                    if sheet['K12'].value != None:
                        if type(sheet['K12'].value) == datetime:
                            Primary_address.Start_Date = sheet['K12'].value
                        else:
                            Primary_address.Start_Date = None

                    Primary_address.End_Date = None
                    Primary_address.Current_Flag = 1
                    
                    Primary_address.Verify_Date = Primary_verification_date
                    Primary_address.Notes = Primary_verification_Notes

                    if Primary_address.Apt_No != 'None':
                        Principle_Residence = Primary_address.Apt_No + ', ' + Primary_address.Street_No + ' ' + Primary_address.Street_Name + ', ' + Primary_address.City + ', ' + Primary_address.Province + ', ' + Primary_address.Country
                    else:
                        Principle_Residence = Primary_address.Street_No + ' ' + Primary_address.Street_Name + ', ' + Primary_address.City + ', ' + Primary_address.Province + ', ' + Primary_address.Country

                    result=db.SaveAddress(Primary_address)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                    
                else:
                    Principle_Residence = None

                # Save Primary applicant's previous address
            
                if sheet['A14'].value != None:
                    Primary_address = Address() 

                    Primary_address.PersonID = Primary_applicant.PersonID
                    Primary_address.Apt_No = str(sheet['E14'].value)
                    Primary_address.Street_No = str(sheet['A14'].value)
                    Primary_address.Street_Name = sheet['B14'].value
                    Primary_address.City = sheet['F14'].value
                    Primary_address.Province = sheet['H14'].value
                    Primary_address.Country = "Canada"
                    Primary_address.Postcode = sheet['I14'].value
                    Primary_address.Living_Status = None
                    
                    if sheet['K14'].value != None:
                        if type(sheet['K14'].value) == datetime:
                            Primary_address.Start_Date = sheet['K14'].value
                        else:
                            Primary_address.Start_Date = None
                    
                    if sheet['K12'].value != None:
                        if type(sheet['K12'].value) == datetime:
                            Primary_address.End_Date = sheet['K12'].value
                        else:
                            Primary_address.End_Date = None
                    
                    Primary_address.Current_Flag = 0
                    Primary_address.Verify_Date = Primary_verification_date
                    Primary_address.Notes = Primary_verification_Notes
                    
                    result=db.SaveAddress(Primary_address)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                # Save SIN
                if Primary_applicant.SIN != None:
                    Primary_SIN = ID()
                    Primary_SIN.PersonID = Primary_applicant.PersonID
                    Primary_SIN.ID_Type = 'SIN'
                    Primary_SIN.ID_Number = Primary_applicant.SIN
                    Primary_SIN.Issue_Date = None
                    Primary_SIN.Expiry_Date = None
                    Primary_SIN.Issue_Country = 'Canada'
                    Primary_SIN.Issue_Province = None
                    Primary_SIN.Start_Date = None
                    Primary_SIN.End_Date = None
                    Primary_SIN.Current_Flag = 1
                    Primary_SIN.Verify_Date = Primary_verification_date
                    Primary_SIN.Notes = Primary_verification_Notes

                    result=db.SaveID(Primary_SIN)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                # Save primary ID
                if sheet['A18'].value != None:
                    Primary_ID = ID()
                    Primary_ID.PersonID = Primary_applicant.PersonID
                    Primary_ID.ID_Type = sheet['A18'].value
                    Primary_ID.ID_Number = sheet['C18'].value
                    Primary_ID.Issue_Date = sheet['F18'].value
                    Primary_ID.Expiry_Date = sheet['H18'].value
                    Primary_ID.Issue_Country = 'Canada'
                    Primary_ID.Issue_Province = sheet['D19'].value
                    Primary_ID.Start_Date = None
                    Primary_ID.End_Date = None
                    Primary_ID.Current_Flag = 1
                    Primary_ID.Verify_Date = Primary_verification_date
                    Primary_ID.Notes = Primary_verification_Notes

                    result=db.SaveID(Primary_ID)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                # Save Primary current employment
                if sheet['C27'].value != None:
                    Primary_Employment = Employment()
                    Primary_Employment.PersonID = Primary_applicant.PersonID
                    Primary_Employment.Employment_Status = sheet['C27'].value
                    Primary_Employment.Employer = sheet['C29'].value
                    Primary_Employment.Industry = sheet['C33'].value
                    Primary_Employment.Occupation = sheet['C34'].value
                    Primary_Employment.Income = sheet['C28'].value
                    Primary_Employment.Unit = sheet['C31'].value
                    Primary_Employment.Street_No = sheet['C30'].value
                    Primary_Employment.Street_Name = sheet['D30'].value
                    Primary_Employment.City = sheet['D31'].value
                    Primary_Employment.Province = sheet['E31'].value
                    Primary_Employment.Country = None
                    Primary_Employment.Postcode = sheet['F31'].value
                    Primary_Employment.Workphone = sheet['K16'].value

                    if sheet['D32'] == None:
                        Primary_Employment.Start_Date = Application_date - timedelta(days=732)
                    else:
                        Primary_Employment.Start_Date = sheet['D32'].value
                    
                    if type(sheet['F32'].value) == datetime:
                        Primary_Employment.End_Date = sheet['F32'].value
                    else:
                        if sheet['F32'].value == None or "present".lower() in sheet['F32'].value.lower() or "now".lower() in sheet['F32'].value.lower() or "current".lower() in sheet['F32'].value.lower():
                            if Primary_Employment.Start_Date != None:
                                Primary_Employment.End_Date = Application_date
                            else:
                                Primary_Employment.End_Date = None
                        
                    Primary_Employment.Current_Flag = 1
                    Primary_Employment.Verify_Date = Primary_verification_date
                    Primary_Employment.Notes = Primary_verification_Notes
                    
                    result=db.SaveEmployment(Primary_Employment)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                # Save Primary Previous employment
                if sheet['G27'].value != None:
                    Primary_Employment = Employment()
                    Primary_Employment.PersonID = Primary_applicant.PersonID,
                    Primary_Employment.Employment_Status = sheet['G27'].value,
                    Primary_Employment.Employer = sheet['G29'].value,
                    Primary_Employment.Industry = sheet['G33'].value
                    Primary_Employment.Occupation = sheet['G34'].value,
                    Primary_Employment.Income = sheet['G28'].value,
                    Primary_Employment.Unit = sheet['G31'].value,
                    Primary_Employment.Street_No = sheet['G30'].value,
                    Primary_Employment.Street_Name = sheet['H30'].value,
                    Primary_Employment.City = sheet['H31'].value,
                    Primary_Employment.Province = sheet['J31'].value,
                    Primary_Employment.Country = None,
                    Primary_Employment.Postcode = sheet['L31'].value,
                    Primary_Employment.Workphone = None,
                    
                    if sheet['H32'] == None:
                        Primary_Employment.Start_Date = Application_date - timedelta(days=732)
                    else:
                        Primary_Employment.Start_Date = sheet['H32'].value
                    
                    if type(sheet['L32'].value) == datetime:
                        Primary_Employment.End_Date = sheet['L32'].value
                    else:
                        if sheet['L32'].value == None or "present".lower() in sheet['L32'].value.lower() or "now".lower() in sheet['L32'].value.lower() or "current".lower() in sheet['L32'].value.lower():
                            if Primary_Employment.Start_Date != None:
                                Primary_Employment.End_Date = Application_date
                            else:
                                Primary_Employment.End_Date = None
                    
                    Primary_Employment.Current_Flag = 0,
                    Primary_Employment.Verify_Date = Primary_verification_date,
                    Primary_Employment.Notes = Primary_verification_Notes
                        
                    result=db.SaveEmployment(Primary_Employment)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                
                # The beginning of Spouse
                if sheet['C22'].value != None and sheet['G22'].value != None:
                    s = Spouse()
                    s.Applicant_PID = Primary_applicant.PersonID
                    s.PersonID = create_personID()
                    s.First_Name = sheet['C22'].value
                    s.Last_Name = sheet['G22'].value
                    s.AdvisorPID = Primary_applicant.AdvisorPID
                    s.Date_of_Birth = sheet['K22'].value
                    
                    result=db.SaveSpouse(s)
                    
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
            
                # The end of Spouse

                # The beginning of Benificiary
                for i in range(5):
                    if sheet.cell(row=i+37,column=2).value != None and sheet.cell(row=i+37,column=4).value != None:
                        B = Beneficiary()
                        B.Applicant_PID = Primary_applicant.PersonID
                        B.PersonID = create_personID()
                        B.First_Name = sheet.cell(row=i+37,column=4).value
                        B.Last_Name = sheet.cell(row=i+37,column=2).value
                        B.B_Relationship=sheet.cell(row=i+37,column=6).value
                        B.Gender = sheet.cell(row=i+37,column=8).value
                        B.Revokable = sheet.cell(row=i+37,column=9).value
                        B.Date_of_Birth = sheet.cell(row=i+37,column=10).value
                        if sheet.cell(row=i+37,column=12).value != None:
                            B.B_Percentage = sheet.cell(row=i+37,column=12).value * 100
                        SINNo = list(str(sheet.cell(row=i+37,column=13).value))
                        if len(SINNo) == 9:
                            if Validate_SIN(SINNo):
                                B.SIN = str(sheet.cell(row=i+37,column=13).value)
                            else:
                                print(B.Full_Name + ' ' + str(sheet.cell(row=i+37, column=13).value) + ': The SIN is NOT valid')
                        else:
                            B.SIN = None
                            B.AdvisorPID = Primary_applicant.AdvisorPID
                        result=db.SaveBeneficiary(B)
                        
                        print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                        

                # Save Trustee
                if sheet['B42'].value != None and sheet['D42'].value != None:
                    T = Trustee()
                    T.Applicant_PID = Primary_applicant.PersonID
                    T.PersonID = create_personID()
                    T.First_Name = sheet['D42'].value
                    T.Last_Name = sheet['B42'].value
                    T.T_Relationship = sheet['F42'].value
                    T.AdvisorPID = Primary_applicant.AdvisorPID
                    
                    result=db.SaveTrustee(T)
                    
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                    
                # Save Co_applicant
                if sheet['L6'].value != None:
                    if sheet['L6'].value.lower() == "yes".lower():
                        Co_Applicant_exist = True
                        sheet=workbook['Co Applicant Info']
                        Co_verification_date = sheet['J18'].value
                        Co_verification_Notes = sheet['H19'].value

                        Co_Applicant = Person()
                        Co_Applicant.PersonID = create_personID()
                        Co_Applicant.First_Name = sheet['E7'].value
                        Co_Applicant.Last_Name=sheet['B7'].value
                       
                        if Co_Applicant.First_Name != None and Co_Applicant.Last_Name != None:
                            Co_Applicant.English_Name=sheet['J7'].value
                            Co_Applicant.Gender=sheet['B8'].value
                            
                            D_of_Birth=sheet['H8'].value
                            M_of_Birth=sheet['F8'].value
                            Y_of_Birth=sheet['D8'].value
                            if Y_of_Birth != None and M_of_Birth!= None and D_of_Birth!= None: # Do nothing if there is no Date of Birth
                                Co_Applicant.Date_of_Birth=Date_of_Birth(Y_of_Birth,M_of_Birth,D_of_Birth)
    
                            place_of_birth = sheet['E10'].value
                            if place_of_birth is None:
                                Country_of_Birth = ""
                                Province_of_Birth = ""
                            else:
                                Co_Applicant.Country_of_Birth=CN_Prov_of_birth(place_of_birth)[0]
                                Co_Applicant.Province_of_Birth=CN_Prov_of_birth(place_of_birth)[1]
                            
                            Co_Applicant.Citizenship=sheet['G9'].value
                            Co_Applicant.Tax_Status=sheet['K9'].value
                            
                            if sheet['I10'].value=="Birth":
                                Co_Applicant.Live_in_Canada_Since=Co_Applicant.Date_of_Birth
                            elif sheet['I10'].value == "Date (mm/yyyy)" :
                                Co_Applicant.Live_in_Canada_Since=sheet['L10'].value
                            else: Co_Applicant.Live_in_Canada_Since = None
                            
                            Co_Applicant.Marital_Status=sheet['D9'].value
                            Co_Applicant.Cellphone=sheet['F16'].value
                            Co_Applicant.Homephone=sheet['H16'].value
                            Co_Applicant.Workphone=sheet['K16'].value
                            
                            SINNo = list(str(sheet['K8'].value))
                            if len(SINNo) == 9:
                                if Validate_SIN(SINNo):
                                    Co_Applicant.SIN = str(sheet['K8'].value)
                                else:
                                    print(Co_Applicant.Full_Name + ' ' + str(sheet['K8'].value) + ': The SIN is NOT valid')
                                    Co_Applicant.SIN = str(sheet['K8'].value)    ##for test purpose only  ''  
                            else:
                                Co_Applicant.SIN = None
                                        
                            Co_Applicant.Email=sheet['A16'].value
                            Co_Applicant.Postcode=sheet['I12'].value
                            Co_Applicant.PersonType="" 
                            Co_Applicant.AdvisorPID = Primary_applicant.AdvisorPID 
                            
                            if sheet['E20'].value == "Yes":
                                Co_Applicant.Bankruptcy = 1
                                Co_Applicant.Discharge_Date = sheet['J20'].value
                            else:
                                Co_Applicant.Bankruptcy = 0
                                Co_Applicant.Discharge_Date = None

                            Co_Applicant.Start_date = None 
                            Co_Applicant.End_date = None
                            Co_Applicant.Current_Flag = 1
                            Co_Applicant.Verify_date = Co_verification_date
                            Co_Applicant.Notes = Co_verification_Notes

                            
                            result=db.SaveApplicant(Co_Applicant)
                            
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                            # Save Co_applicant's current address
                            if sheet['B12'].value != None:
                                Co_address = Address() 

                                Co_address.PersonID = Co_Applicant.PersonID
                                Co_address.Apt_No = str(sheet['E12'].value)
                                Co_address.Street_No = str(sheet['A12'].value)
                                Co_address.Street_Name = sheet['B12'].value
                                Co_address.City = sheet['F12'].value
                                Co_address.Province = sheet['H12'].value
                                Co_address.Country = "Canada"
                                Co_address.Postcode = sheet['I12'].value
                                Co_address.Homephone = sheet['H16'].value
                                Co_address.Living_Status = sheet['B9'].value
                                if sheet['K12'].value != None:
                                    if type(sheet['K12'].value) == datetime:
                                        Co_address.Start_Date = sheet['K12'].value
                                    else:
                                        Co_address.Start_Date = None
                                
                                Co_address.End_Date = None
                                Co_address.Current_Flag = 1
                                Co_address.Verify_Date = Co_verification_date
                                Co_address.Notes = Co_verification_Notes

                                result=db.SaveAddress(Co_address)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode) 

                            # Save Co_applicant's previous address
                            if sheet['A14'].value != None:
                                Co_address = Address() 

                                Co_address.PersonID = Co_Applicant.PersonID
                                Co_address.Apt_No = str(sheet['E14'].value)
                                if Co_address.Apt_No!= 'None':
                                    if len(Co_address.Apt_No) > 10:
                                        print()
                                Co_address.Street_No = str(sheet['A14'].value)
                                Co_address.Street_Name = sheet['B14'].value
                                Co_address.City = sheet['F14'].value
                                Co_address.Province = sheet['H14'].value
                                Co_address.Country = "Canada"
                                Co_address.Postcode = sheet['I14'].value
                                Co_address.Homephone = sheet['H16'].value
                                Co_address.Living_Status = None
                                if sheet['K14'].value != None:
                                    if type(sheet['K14'].value) == datetime:
                                        Co_address.Start_Date = sheet['K14'].value
                                    else:
                                        Co_address.Start_Date = None
                                
                                if sheet['K12'].value != None:
                                    if type(sheet['K12'].value) == datetime:
                                        Co_address.End_Date = sheet['K12'].value
                                    else:
                                        Co_address.End_Date = None
                                Co_address.Current_Flag = 0
                                Co_address.Verify_Date = Co_verification_date
                                Co_address.Notes = Co_verification_Notes

                                result=db.SaveAddress(Co_address)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)       
                            
                            # Save Co_applicant's SIN
                            if Co_Applicant.SIN != None:
                                Co_SIN = ID()
                                Co_SIN.PersonID = Co_Applicant.PersonID,
                                Co_SIN.ID_Type = 'SIN',
                                Co_SIN.ID_Number = Co_Applicant.SIN,
                                Co_SIN.Issue_Date = None
                                Co_SIN.Expiry_Date = None,
                                Co_SIN.Issue_Country = 'Canada',
                                Co_SIN.Issue_Province = None,
                                Co_SIN.Current_Flag = 1,
                                Co_SIN.Verify_Date = Co_verification_date,
                                Co_SIN.Notes = Co_verification_Notes

                                result=db.SaveID(Co_SIN)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                            
                            # Save Co_applicant's ID
                            if sheet['A18'].value != None:
                                Co_ID = ID()
                                Co_ID.PersonID = Primary_applicant.PersonID,
                                Co_ID.ID_Type = sheet['A18'].value,
                                Co_ID.ID_Number = sheet['C18'].value,
                                Co_ID.Issue_Date = sheet['F18'].value
                                Co_ID.Expiry_Date = sheet['H18'].value,
                                Co_ID.Issue_Country = 'Canada',
                                Co_ID.Issue_Province = sheet['D19'].value,
                                Co_ID.Current_Flag = 1,
                                Co_ID.Verify_Date = Co_verification_date,
                                Co_ID.Notes = Co_verification_Notes

                                result=db.SaveID(Co_ID)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)


                            # Save co_applicant current employment
                            if sheet['C27'].value != None:
                                Co_Employment = Employment()
                                Co_Employment.PersonID = Primary_applicant.PersonID,
                                Co_Employment.Employment_Status = sheet['C27'].value,
                                Co_Employment.Employer = sheet['C29'].value,
                                Co_Employment.Industry = sheet['C33'].value
                                Co_Employment.Occupation = sheet['C34'].value,
                                Co_Employment.Income = sheet['C28'].value,
                                Co_Employment.Unit = sheet['C31'].value,
                                Co_Employment.Street_No = sheet['C30'].value,
                                Co_Employment.Street_Name = sheet['D30'].value,
                                Co_Employment.City = sheet['D31'].value,
                                Co_Employment.Province = sheet['E31'].value,
                                Co_Employment.Country = None,
                                Co_Employment.Postcode = sheet['F31'].value
                                
                                if sheet['D32'] == None:
                                    Co_Employment.Start_Date = Application_date - timedelta(days=732)
                                else:
                                    Co_Employment.Start_Date = sheet['D32'].value
                                
                                if type(sheet['F32'].value) == datetime:
                                    Co_Employment.End_Date = sheet['F32'].value
                                else:
                                    if sheet['F32'].value == None or "present".lower() in sheet['F32'].value.lower() or "now".lower() in sheet['F32'].value.lower() or "current".lower() in sheet['F32'].value.lower():
                                        if Co_Employment.Start_Date != None:
                                            Co_Employment.End_Date = Application_date
                                        else:
                                            Co_Employment.End_Date = None
                                 
                                Co_Employment.Current_Flag = 1,
                                Co_Employment.Verify_Date = Co_verification_date,
                                Co_Employment.Notes = Co_verification_Notes
                    
                                result=db.SaveEmployment(Co_Employment)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                            # Save co_applicant Previous employment
                            if sheet['G27'].value != None:
                                Co_Employment = Employment()
                                Co_Employment.PersonID = Primary_applicant.PersonID,
                                Co_Employment.Employment_Status = sheet['G27'].value,
                                Co_Employment.Employer = sheet['G29'].value,
                                Co_Employment.Industry = sheet['G33'].value
                                Co_Employment.Occupation = sheet['G34'].value,
                                Co_Employment.Unit = sheet['G31'].value,
                                Co_Employment.Street_No = sheet['G30'].value,
                                Co_Employment.Street_Name = sheet['H30'].value,
                                Co_Employment.City = sheet['H31'].value,
                                Co_Employment.Province = sheet['J31'].value,
                                Co_Employment.Postcode = sheet['L31'].value
                                Co_Employment.Income = sheet['G28'].value,
                                
                                if sheet['H32'] == None:
                                    Co_Employment.Start_Date = Application_date - timedelta(days=732)
                                else:
                                    Co_Employment.Start_Date = sheet['H32'].value
                                
                                if type(sheet['L32'].value) == datetime:
                                    Co_Employment.End_Date = sheet['L32'].value
                                else:
                                    if sheet['L32'].value == None or "present".lower() in sheet['L32'].value.lower() or "now".lower() in sheet['L32'].value.lower() or "current".lower() in sheet['L32'].value.lower():
                                        if Co_Employment.Start_Date != None:
                                            Co_Employment.End_Date = Application_date
                                        else:
                                            Co_Employment.End_Date = None

                                Co_Employment.Current_Flag = 0,
                                Co_Employment.Verify_Date = Co_verification_date,
                                Co_Employment.Notes = Co_verification_Notes

                                result=db.SaveEmployment(Co_Employment)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                        else:
                            Co_Applicant_exist = False
                    else:
                        Co_Applicant_exist = False        
                        
                    
                
                # Save Primary applicant's Assets and Liability
                sheet=workbook['Assets & Liabilities']
                real_estate_num = 43
                for i in range(5,20):
                    if sheet.cell(row=i,column=1).value != None and sheet.cell(row=i,column=1).value != "New Apply":
                        if sheet.cell(row=i,column=2).value != None or sheet.cell(row=i,column=3).value != None:
                            if i == 5:
                                new_address = Principle_Residence
                            elif "Real Estate" in sheet.cell(row=i,column=1).value:
                                new_address = sheet.cell(row = real_estate_num,column=2).value
                                real_estate_num = real_estate_num + 1
                        
                        if sheet.cell(row=i,column=2).value != None:
                            
                            new_asset = Asset()
                            
                            new_asset.PersonID = Primary_applicant.PersonID
                            new_asset.Assets_Type = sheet.cell(row=i,column=1).value
                            new_asset.Institution = sheet.cell(row=i,column=4).value
                            new_asset.Market_Value = sheet.cell(row=i,column=2).value
                            new_asset.Verify_Date = Primary_verification_date
                            new_asset.Notes = Primary_verification_Notes
                            new_asset.Address = new_address
                            
                            result=db.SaveAsset(new_asset)
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                        #Save one line of liability that associated with Asset above:
                        if sheet.cell(row=i,column=3).value != None:
                            new_liability = Liability()
                            new_liability.PersonID = Primary_applicant.PersonID
                            new_liability.L_Balance = sheet.cell(row=i,column=3).value
                            new_liability.L_Monthly_Payment = sheet.cell(row=i,column=6).value
                            new_liability.Institution = sheet.cell(row=i,column=4).value
                            new_liability.Verify_Date = Primary_verification_date
                            new_liability.Notes = Primary_verification_Notes

                            if i == 5 or "Real Estate" in sheet.cell(row=i,column=1).value:
                                new_liability.Address = new_address
                                new_liability.L_Type = "Mortgage"
                            else:
                                if sheet.cell(row=i,column=1).value != None or sheet.cell(row=i,column=1).value != "New Apply":
                                    new_liability.L_Type = sheet.cell(row=i,column=1).value

                            result=db.SaveLiability(new_liability)
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                        
                        #Save vehicle as liability:
                        if i == 9 and sheet.cell(row=i,column=6).value != None:
                            new_liability = Liability()

                            new_liability.PersonID = Primary_applicant.PersonID
                            
                            new_liability.L_Monthly_Payment = sheet.cell(row=i,column=6).value
                            new_liability.Institution = sheet.cell(row=i,column=4).value
                            new_liability.Verify_Date = Primary_verification_date
                            new_liability.Notes = Primary_verification_Notes
                            new_liability.L_Type = "Vihicle"
                            
                            result=db.SaveLiability(new_liability)
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                        # Save property tax as one liability
                        if sheet.cell(row=i,column=7).value != None: 
                            new_property_tax = Liability()
                            
                            new_property_tax.PersonID = Primary_applicant.PersonID
                            
                            new_property_tax.L_Monthly_Payment = sheet.cell(row=i,column=7).value
                            new_property_tax.Institution = "City"
                            new_property_tax.Verify_Date = Primary_verification_date
                            new_property_tax.Notes = Primary_verification_Notes
                            new_property_tax.L_Type = "Property Tax"
                            # property tax address should be equal to the new liability above
                            new_property_tax.Address = new_liability.Address

                            result=db.SaveLiability(new_property_tax)
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                        # Save condo fee as one liability
                        if sheet.cell(row=i,column=8).value != None:
                            if sheet.cell(row=i,column=8).value != 0: 
                                new_condo_fee = Liability()
                                
                                new_condo_fee.PersonID = Primary_applicant.PersonID
                                
                                new_condo_fee.L_Monthly_Payment = sheet.cell(row=i,column=8).value
                                new_condo_fee.L_Type = "Condo fee"
                                new_condo_fee.Institution = "Condo Corporation"
                                new_condo_fee.Verify_Date = Primary_verification_date
                                new_condo_fee.Notes = Primary_verification_Notes
                                # property tax address should be equal to the new liability above
                                new_condo_fee.Address = new_liability.Address

                                result=db.SaveLiability(new_condo_fee)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                
                #Save TFSA, RRSP etc asset for both primary applicant and co applicant
                for i in range(23,34):
                    if sheet.cell(row=i,column=1).value != None:
                        #Save TFSA, RRSP etc asset for primary applicant
                        if sheet.cell(row=i,column=2).value != None:
                            
                            new_asset = Asset()
                            
                            new_asset.PersonID = Primary_applicant.PersonID
                            new_asset.Assets_Type = sheet.cell(row=i,column=1).value
                            new_asset.Institution = sheet.cell(row=i,column=4).value
                            new_asset.Market_Value = sheet.cell(row=i,column=2).value
                            new_asset.Verify_Date = Primary_verification_date
                            new_asset.Notes = Primary_verification_Notes
                            
                            
                            result=db.SaveAsset(new_asset)
                            print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                        
                        #Save TFSA, RRSP etc asset for co applicant
                        if sheet.cell(row=i,column=3).value != None:
                            if Co_Applicant_exist == True:
                                new_asset = Asset()
                                
                                new_asset.PersonID = Co_Applicant.PersonID
                                new_asset.Assets_Type = sheet.cell(row=i,column=1).value
                                new_asset.Institution = sheet.cell(row=i,column=4).value
                                new_asset.Market_Value = sheet.cell(row=i,column=3).value
                                new_asset.Verify_Date = Co_verification_date
                                new_asset.Notes = Co_verification_Notes
                                
                                
                                result=db.SaveAsset(new_asset)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                            else:
                                print ("This client has no Co_applicant but has Co_applicant's asset value" + Primary_applicant.Full_Name)

                #Save Income for both primary applicant and co applicant
                
                for i in range(24,29):
                    #Save Income for primary applicant
                    if sheet.cell(row=i,column=7).value != None or sheet.cell(row=i,column=8).value != 0:
                        new_income = Income()
                        
                        new_income.PersonID = Primary_applicant.PersonID
                        new_income.I_Type = sheet.cell(row=i,column=6).value
                        if new_income.I_Type == 'Rental':
                            new_income.I_Frequency = 'Monthly'
                            new_income.Income = sheet.cell(row=i,column=8).value
                        else:
                            new_income.I_Frequency = 'Annual'
                            new_income.Income = sheet.cell(row=i,column=7).value
                        new_income.notes = Primary_verification_Notes 
                        new_income.Current_Flag = 1
                        
                        result=db.SaveIncome(new_income)
                        print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                    
                for i in range(31,36):
                    #Save Income for Co applicant
                    if sheet.cell(row=i,column=7).value != None or sheet.cell(row=i,column=8).value != 0:
                        print(sheet.cell(row=i,column=7).value)
                        print(sheet.cell(row=i,column=8).value)
                        new_income = Income()
                        
                        new_income.PersonID = Co_Applicant.PersonID
                        new_income.I_Type = sheet.cell(row=i,column=6).value
                        if new_income.I_Type == 'Rental':
                            new_income.I_Frequency = 'Monthly'
                            new_income.Income = sheet.cell(row=i,column=8).value
                        else:
                            new_income.I_Frequency = 'Annual'
                            new_income.Income = sheet.cell(row=i,column=7).value
                        new_income.notes = Co_verification_Notes
                        new_income.Current_Flag = 1
                        
                        result=db.SaveIncome(new_income)
                        print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                        
                        #Save TFSA, RRSP etc asset for co applicant
                        if sheet.cell(row=i,column=3).value != None:
                            if Co_Applicant_exist == True:
                                new_asset = Asset()
                                
                                new_asset.PersonID = Co_Applicant.PersonID
                                new_asset.Assets_Type = sheet.cell(row=i,column=1).value
                                new_asset.Institution = sheet.cell(row=i,column=4).value
                                new_asset.Market_Value = sheet.cell(row=i,column=3).value
                                new_asset.Verify_Date = Co_verification_date
                                new_asset.Notes = Co_verification_Notes
                                
                                
                                result=db.SaveAsset(new_asset)
                                print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                            else:
                                print ("This client has no Co_applicant but has Co_applicant's asset value" + Primary_applicant.Full_Name)
                
                # Save Primary applicant's KYC
                K = KYC()
                K.PersonID = Primary_applicant.PersonID
                K.Verify_Date = Primary_verification_date
                K.Notes = Primary_verification_Notes
                sheet_iA=workbook['KYC-iA']
                sheet_ML=workbook['KYC-ML']
                sheet_CL=workbook['KYC-CL']
                if sheet_iA['B4'].value != None and sheet_iA['B4']:
                    score_cell = sheet_iA['B4':'B11'] # iA score cells
                    K.Score = [str(score_cell[x][0].value) for x in range(len(score_cell))] 
                    K.VersionNo = "IA_V1"
                    result=db.SaveKYC(K)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                if sheet_ML['B4'].value != None and sheet_ML['B4']:
                    score_cell = sheet_ML['B4':'B11'] # iA score cells
                    K.Score = [str(score_cell[x][0].value) for x in range(len(score_cell))]
                    K.VersionNo = "ML_V1"
                    result=db.SaveKYC(K)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)
                if sheet_CL['B4'].value != None and sheet_CL['B4']:                    
                    score_cell = sheet_CL['B4':'B21']
                    K.Score = [str(score_cell[x][0].value) for x in range(len(score_cell))]
                    section_subtotal = [2,6,10,15] #section subtal of CL spreadsheet
                    for x in reversed(section_subtotal): K.Score.pop(x)
                    K.VersionNo = "CL_V1"
                    result=db.SaveKYC(K)
                    print(result.ErrorMsg,result.Succeeded,result.ErrorCode)

                    
                    
                



# create the list of PersonID
def create_personID():
    global dateTimeObj, PersonID_index
    PersonID_index += 1
    
    return (dateTimeObj + timedelta(seconds=PersonID_index)).strftime("EON0000%Y%m%d%H%M%S")

dateTimeObj = datetime.now()

print(dateTimeObj)

PersonID_index = 0


# To locate Workbook
#list all files under excel directory and get each file
## Set path
my_project_dir = pathlib.Path('/Users/telus/dev')
my_excels_dir = my_project_dir.joinpath('Excels')
my_scripts_dir = my_project_dir.joinpath('Python_Scripts')
entries = os.listdir(my_excels_dir)
if '.DS_Store' in entries:
    entries.remove('.DS_Store')

main_application(entries)
#maxThread = 1 
#num_of_files = len(entries)
#threadRange =  int(math.modf(num_of_files / maxThread)[1])
#threadList = []

#try:

#    for i in range(maxThread):        
#        thread = threading.Thread(target=main_application, args=(entries[i*threadRange: (num_of_files if i == maxThread - 1 else (i+1)*threadRange)],))        
#        thread.start()
#        threadList.append(thread)

#    for i in range(maxThread):
#        threadList[i].join()
#except Exception as e:
#   print ("Error: unable to start thread")
#print(datetime.now()-dateTimeObj)